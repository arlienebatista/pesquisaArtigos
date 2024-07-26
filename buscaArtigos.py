import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
import webbrowser
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from datetime import datetime

# Função para buscar artigos na API CrossRef
def search_articles():
    query = entry.get()  # Obtém o termo de busca da entrada do usuário
    if not query:
        messagebox.showwarning("Ops!", "Por favor, insira um termo de busca")
        return

    url = "https://api.crossref.org/works"
    params = {"query.title": query, "rows": 100}  # Limita os resultados e busca pelo título

    try:
        response = requests.get(url, params=params)
        response.raise_for_status()  # Lança uma exceção para erros HTTP
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Erro de Requisição", str(e))
        return

    results = response.json().get('message', {}).get('items', [])
    if not results:
        messagebox.showinfo("Nenhum Resultado", "Nenhum artigo encontrado para o termo de busca")
        return

    # Função interna para obter o ano de um artigo
    def get_year(item):
        date_parts = item.get("issued", {}).get("date-parts", [[None]])[0]
        return date_parts[0] if date_parts[0] is not None else 0

    # Ordena os resultados pelo ano em ordem decrescente
    results = sorted(results, key=get_year, reverse=True)

    # Limpa os itens existentes na árvore antes de adicionar novos resultados
    for i in tree.get_children():
        tree.delete(i)

    # Adiciona os resultados na árvore (Treeview)
    for item in results:
        title = item.get("title", ["Sem Título"])[0]  # Obtém o título do artigo
        author = ", ".join([author.get("given", "") + " " + author.get("family", "") for author in item.get("author", [])])  # Obtém os autores
        year = get_year(item)  # Obtém o ano
        link = item.get("URL", "Sem URL")  # Obtém o link

        # Insere uma nova linha na árvore com os dados do artigo
        tree.insert("", "end", values=(title, author, year, link))

# Função para abrir o link do artigo ao clicar duas vezes na árvore
def open_link(event):
    selected_item = tree.selection()[0]  # Obtém o item selecionado
    link = tree.item(selected_item, "values")[3]  # Obtém o link do item
    if link and link != "Sem URL":
        webbrowser.open(link)  # Abre o link no navegador padrão
    else:
        messagebox.showinfo("Sem URL", "Este artigo não possui um URL")

# Função para limpar a busca e os resultados exibidos
def clear_search():
    entry.delete(0, tk.END)  # Limpa o campo de entrada
    for i in tree.get_children():
        tree.delete(i)  # Remove todos os itens da árvore

# Função para salvar os resultados em um arquivo Excel
def save_results():
    # Abre a caixa de diálogo para selecionar o local e nome do arquivo
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")])
    if save_path:
        wb = Workbook()  # Cria uma nova planilha do Excel
        ws = wb.active
        ws.title = "Resultados da Busca"  # Define o título da planilha
        
        # Adiciona os títulos das colunas na primeira linha
        ws.append(["Título", "Autor", "Ano", "Link"])

        # Define o estilo de ano para as células
        year_style = NamedStyle(name='year_style', number_format='YYYY')
        
        for row in tree.get_children():
            values = tree.item(row, 'values')  # Obtém os valores do item
            title, author, year, link = values
            
            # Converte o ano para uma data com o primeiro dia do ano, ou deixa em branco se o ano for 0
            if year != 0:
                try:
                    year_date = datetime(year=int(year), month=1, day=1)
                except ValueError:
                    year_date = ""
            else:
                year_date = ""
            
            values = (title, author, year_date, link)
            
            # Adiciona os valores na planilha
            ws.append(values)
            
            # Aplica o estilo de ano à célula correspondente, se o ano não for 0
            if year != 0:
                ws.cell(row=ws.max_row, column=3).style = year_style

        try:
            wb.save(save_path)  # Salva a planilha no caminho especificado
            messagebox.showinfo("Salvo com Sucesso", f"Resultados salvos em {save_path}")
        except Exception as e:
            messagebox.showerror("Erro ao Salvar", f"Ocorreu um erro ao salvar o arquivo: {str(e)}")

# Função para fechar o programa
def close_program():
    root.destroy()  # Fecha a janela principal e encerra o programa

# Cria a janela principal da aplicação
root = tk.Tk()
root.title("Busca de Artigos")  # Define o título da janela
root.geometry("950x600")  # Define o tamanho da janela
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
root.rowconfigure(1, weight=0)
root.rowconfigure(2, weight=0)

# Cria o frame principal para os widgets
frame = ttk.Frame(root, padding="10")
frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
frame.columnconfigure(0, weight=1)
frame.rowconfigure(2, weight=1)

# Adiciona o rótulo informativo acima da barra de busca
info_label = ttk.Label(frame, text="Para obter os resultados basta informar o termo que deseja encontrar e clicar em 'Buscar'. A aplicação usa API gratuita CrossRef.", wraplength=800)
info_label.grid(row=0, column=0, columnspan=4, padx=5, pady=5, sticky=tk.W)

# Cria o rótulo para o campo de busca
search_label = ttk.Label(frame, text="Digite o termo:")
search_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)

# Cria a entrada de texto para o termo de busca
entry = ttk.Entry(frame, width=100)
entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)

# Cria o botão de busca
search_button = ttk.Button(frame, text="Buscar", command=search_articles)
search_button.grid(row=1, column=2, padx=5, pady=5, sticky=tk.W)

# Cria o botão de limpar
clear_button = ttk.Button(frame, text="Limpar", command=clear_search)
clear_button.grid(row=1, column=3, padx=5, pady=5, sticky=tk.W)

# Define as colunas da árvore de resultados
columns = ("Title", "Author", "Year", "Link")
tree = ttk.Treeview(frame, columns=columns, show="headings")
tree.heading("Title", text="Título")
tree.heading("Author", text="Autor")
tree.heading("Year", text="Ano")
tree.heading("Link", text="Link")

tree.column("Title", width=250)
tree.column("Author", width=160)
tree.column("Year", width=30)
tree.column("Link", width=250)

tree.grid(row=2, column=0, columnspan=4, sticky=(tk.N, tk.S, tk.W, tk.E))

# Adiciona uma barra de rolagem vertical à árvore
scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
tree.configure(yscroll=scrollbar.set)
scrollbar.grid(row=2, column=4, sticky=(tk.N, tk.S))

# Configura a ação de abrir o link ao dar um duplo clique na árvore
tree.bind("<Double-1>", open_link)

# Cria o frame para os botões de controle
button_frame = ttk.Frame(root, padding="10")
button_frame.grid(row=1, column=0, sticky=(tk.E, tk.W))
button_frame.columnconfigure(0, weight=1)  # Expande a coluna 0 para preencher o espaço horizontal
button_frame.columnconfigure(1, weight=1)  # Expande a coluna 1 para preencher o espaço horizontal

# Cria o botão de salvar
save_button = ttk.Button(button_frame, text="Salvar", command=save_results)
save_button.grid(row=0, column=0, padx=5, pady=5, sticky=tk.E)  # Centraliza o botão à direita

# Cria o botão de fechar
close_button = ttk.Button(button_frame, text="Fechar", command=close_program)
close_button.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)  # Centraliza o botão à esquerda

# Adiciona o rodapé com informações de autoria e ano
footer_label = ttk.Label(root, text="2024 © Desenvolvido por Arliene Santos ", anchor=tk.CENTER)
footer_label.grid(row=2, column=0, sticky=(tk.W, tk.E), padx=5, pady=5)

# Inicia o loop principal da interface gráfica
root.mainloop()
